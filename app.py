import os
import uuid
import openpyxl
import pickle
import seaborn as sns
import matplotlib.pyplot as plt
from flask import Flask, render_template, request, jsonify
from mlxtend.frequent_patterns import fpgrowth
from mlxtend.preprocessing import TransactionEncoder
import pandas as pd
from helpers import (generate_conditional_pattern_base,
                     build_conditional_fp_tree,
                     generate_frequent_patterns,
                     generate_frequent_2_itemsets,
                     evaluate_association_rules,
                     get_item_names)
from collections import Counter
import matplotlib
matplotlib.use('agg')

# Load model
with open('Model.pkl', 'rb') as model_file:
    frequent_patterns_result = pickle.load(model_file)

app = Flask(__name__)
te = TransactionEncoder()

#Lokasi setelah sudah upload
app.config['UPLOAD_FOLDER'] = 'static/uploads/'

@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        file = request.files['file']
        if file and file.filename.endswith('.xlsx'):
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
            file.save(file_path)

            df = pd.read_excel(file_path)
            transactions = df.groupby('order no').apply(
                lambda x: list(x['item name'])).tolist()
            
            te_ary = te.fit(transactions).transform(transactions)
            df_encoded = pd.DataFrame(te_ary, columns=te.columns_)
            frequent_itemsets = fpgrowth(
                df_encoded, min_support=0.01, use_colnames=True)
            frequent_itemsets = frequent_itemsets.sort_values(
                by='support', ascending=False)

            # Membuat FP-Tree berdasarkan frequent itemsets
            fp_tree = {}
            for index, row in frequent_itemsets.iterrows():
                current_node = fp_tree
                for item in row['itemsets']:
                    if item not in current_node:
                        current_node[item] = {'count': row['support'], 'children': {}}
                    else:
                        current_node[item]['count'] += row['support']
                    current_node = current_node[item]['children']

            prefix_example = []
            conditional_pattern_base_result = generate_conditional_pattern_base(fp_tree, prefix_example)

            conditional_fp_tree_result = build_conditional_fp_tree(conditional_pattern_base_result)

            # Menggunakan Conditional FP-tree untuk pembangkitan Frequent Patterns
            min_support_threshold = 0.01
            frequent_patterns_result = {}
            generate_frequent_patterns(conditional_fp_tree_result, min_support_threshold, [], frequent_patterns_result)

            # Menggunakan fungsi untuk menghasilkan Frequent 2-itemsets
            min_support_2_itemsets = 0.01
            frequent_2_itemsets_result = {}
            generate_frequent_2_itemsets(conditional_fp_tree_result, min_support_2_itemsets, frequent_2_itemsets_result)

            # Mencari Support 2 Itemset
            support_2_itemset_result = {}
            total_transactions = len(transactions)

            for itemset, support in frequent_2_itemsets_result.items():
                support_2_itemset_result[itemset] = support / total_transactions

            # Mencari Confidence 2 Itemset
            confidence_2_itemset_result = {}

            for itemset, support in frequent_2_itemsets_result.items():
                item_A, item_B = itemset
                support_A = frequent_patterns_result.get((item_A,), 0)
                confidence = support / support_A
                confidence_2_itemset_result[itemset] = confidence

            evaluation_results = evaluate_association_rules(
                frequent_2_itemsets_result, frequent_patterns_result)

            categories = df["item group"].unique()

            # Menghitung total count dan support untuk setiap produk
            transaction_counters = [Counter(transaction) for transaction in transactions]
            product_counts = Counter()
            for counter in transaction_counters:
                product_counts.update(counter)

            product_supports = {product: count / total_transactions for product, count in product_counts.items()}

            product_freq_df = pd.DataFrame.from_dict(
                product_counts, orient='index', columns=['Total Count']).sort_values(by="Total Count", ascending=False)
            product_freq_df = product_freq_df.reset_index().rename(columns={'index': 'Product'})
            product_freq_df['Support'] = product_freq_df['Total Count'] / total_transactions

            plt.figure(figsize=(10, 8))
            barplot = sns.barplot(x='Total Count', y='Product',
                                data=product_freq_df.head(10), palette='Blues_d')
            plt.title('Top 10 Product Frequencies')
            plt.xlabel('Total Count')
            plt.ylabel('Product')

            # Menambahkan anotasi frekuensi pada setiap bar
            for index, row in product_freq_df.head(10).iterrows():
                barplot.text(row['Total Count'], index, f"{row['Total Count']}",
                            va='center', fontsize=10, bbox=dict(facecolor='white', alpha=0.5))

            sns.despine(left=True, bottom=True)

            frequency_plot_path = "static/product_frequency_plot.png"
            plt.savefig(frequency_plot_path, bbox_inches='tight')

            return render_template("pages/index.html",
                                evaluation_results=evaluation_results,
                                categories=categories,
                                product_counts=product_counts,
                                product_supports=product_supports,
                                frequency_plot_path=frequency_plot_path,
                                result=True,
                                )
    else:
        return render_template("pages/index.html")

if __name__ == "__main__":
    app.run(port=8080)




# @app.route("/get_item_names", methods=["POST"])
# def get_item_names_route():
#     df = pd.read_excel("BulanMei2022.xlsx")
#     category = request.form.get("category")
#     item_names = get_item_names(df, category)
#     return jsonify(item_names)

# @app.route('/save_itemsets', methods=['POST'])
# def save_itemsets():
#     data = request.json
#     categories = data.get('categories', [])
#     products = data.get('products', [])

#     try:
#         # Load file Excel yang sudah ada
#         workbook = openpyxl.load_workbook("BulanMei2022.xlsx")
#         sheet = workbook.active

#         # Ambil nomor baris terakhir
#         last_row = sheet.max_row

#         # Tambahkan data baru ke baris selanjutnya
#         for category, product in zip(categories, products):
#             new_row = last_row + 1
#             sheet.cell(row=new_row, column=1, value=str(uuid.uuid4()))  # order_id
#             sheet.cell(row=new_row, column=17, value=category)  # item group
#             sheet.cell(row=new_row, column=18, value=product)  # item name
#             last_row = new_row

#         # Simpan perubahan ke file Excel
#         workbook.save("BulanMei2022.xlsx")

#         return jsonify({"status": "success"}), 200
#     except Exception as e:
#         return jsonify({"status": "error", "message": str(e)}), 500

# @app.route("/")
# def index():
#     df = pd.read_excel("BulanMei2022.xlsx")
#     transactions = df.groupby('order no').apply(
#         lambda x: list(x['item name'])).tolist()
#     te_ary = te.fit(transactions).transform(transactions)
#     df_encoded = pd.DataFrame(te_ary, columns=te.columns_)
#     frequent_itemsets = fpgrowth(
#         df_encoded, min_support=0.01, use_colnames=True)
#     frequent_itemsets = frequent_itemsets.sort_values(
#         by='support', ascending=False)

#     # Membuat FP-Tree berdasarkan frequent itemsets
#     fp_tree = {}
#     for index, row in frequent_itemsets.iterrows():
#         current_node = fp_tree
#         for item in row['itemsets']:
#             if item not in current_node:
#                 current_node[item] = {'count': row['support'], 'children': {}}
#             else:
#                 current_node[item]['count'] += row['support']
#             current_node = current_node[item]['children']

#     prefix_example = []
#     conditional_pattern_base_result = generate_conditional_pattern_base(fp_tree, prefix_example)

#     conditional_fp_tree_result = build_conditional_fp_tree(conditional_pattern_base_result)

#     # Menggunakan Conditional FP-tree untuk pembangkitan Frequent Patterns
#     min_support_threshold = 0.01
#     frequent_patterns_result = {}
#     generate_frequent_patterns(conditional_fp_tree_result, min_support_threshold, [], frequent_patterns_result)

#     # Menggunakan fungsi untuk menghasilkan Frequent 2-itemsets
#     min_support_2_itemsets = 0.01
#     frequent_2_itemsets_result = {}
#     generate_frequent_2_itemsets(conditional_fp_tree_result, min_support_2_itemsets, frequent_2_itemsets_result)

#     # Mencari Support 2 Itemset
#     support_2_itemset_result = {}
#     total_transactions = len(transactions)

#     for itemset, support in frequent_2_itemsets_result.items():
#         support_2_itemset_result[itemset] = support / total_transactions

#     # Mencari Confidence 2 Itemset
#     confidence_2_itemset_result = {}

#     for itemset, support in frequent_2_itemsets_result.items():
#         item_A, item_B = itemset
#         support_A = frequent_patterns_result.get((item_A,), 0)
#         confidence = support / support_A
#         confidence_2_itemset_result[itemset] = confidence

#     evaluation_results = evaluate_association_rules(
#         frequent_2_itemsets_result, frequent_patterns_result)

#     categories = df["item group"].unique()

#     # Menghitung total count dan support untuk setiap produk
#     transaction_counters = [Counter(transaction) for transaction in transactions]
#     product_counts = Counter()
#     for counter in transaction_counters:
#         product_counts.update(counter)

#     product_supports = {product: count / total_transactions for product, count in product_counts.items()}

#     product_freq_df = pd.DataFrame.from_dict(
#         product_counts, orient='index', columns=['Total Count']).sort_values(by="Total Count", ascending=False)
#     product_freq_df = product_freq_df.reset_index().rename(columns={'index': 'Product'})
#     product_freq_df['Support'] = product_freq_df['Total Count'] / total_transactions

#     plt.figure(figsize=(10, 8))
#     barplot = sns.barplot(x='Total Count', y='Product',
#                           data=product_freq_df.head(10), palette='Blues_d')
#     plt.title('Top 10 Product Frequencies')
#     plt.xlabel('Total Count')
#     plt.ylabel('Product')

#     # Menambahkan anotasi frekuensi pada setiap bar
#     for index, row in product_freq_df.head(10).iterrows():
#         barplot.text(row['Total Count'], index, f"{row['Total Count']}",
#                      va='center', fontsize=10, bbox=dict(facecolor='white', alpha=0.5))

#     sns.despine(left=True, bottom=True)

#     frequency_plot_path = "static/product_frequency_plot.png"
#     plt.savefig(frequency_plot_path, bbox_inches='tight')

#     return render_template("pages/index.html",
#                            evaluation_results=evaluation_results,
#                            categories=categories,
#                            product_counts=product_counts,
#                            product_supports=product_supports,
#                            frequency_plot_path=frequency_plot_path,
#                            )